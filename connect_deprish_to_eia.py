"""
Connect the master unit list with depreciation records from PUCs and FERC1.

This module connects the records from the depreciation data to their
appropirate ids in the EIA mater unit list. The PUC and FERC1 data that have
been compiled into an excel spreadsheet. The master unit list is a compilation
of various slices of EIA records.
"""

import argparse
import logging
import pathlib
import sys

import coloredlogs
import pandas as pd
from fuzzywuzzy import fuzz, process
from openpyxl import load_workbook
from xlrd import XLRDError

import make_plant_parts_eia
import pudl

logger = logging.getLogger(__name__)
log_format = '%(asctime)s [%(levelname)8s] %(name)s:%(lineno)s %(message)s'
coloredlogs.install(fmt=log_format, level='INFO', logger=logger)

STRINGS_TO_CLEAN = {
    "combined cycle": ["CC"],
    "combustion turbine": ["CT"],
}

INT_IDS = ['utility_id_ferc1', 'utility_id_pudl',
           'plant_id_pudl', 'report_year']

RESTRICT_MATCH_COLS = ['plant_id_pudl', 'utility_id_pudl', 'report_year']

MUL_COLS = [
    'record_id_eia', 'plant_name_new', 'plant_part', 'report_year',
    'ownership', 'plant_name_eia', 'plant_id_eia', 'generator_id',
    'unit_id_pudl', 'prime_mover_code', 'energy_source_code_1',
    'technology_description', 'ferc_acct_name', 'utility_id_eia',
    'true_gran', 'appro_part_label', 'appro_record_id_eia',
]

MUL_RENAME = {
    'record_id_eia': 'record_id_eia_name_match',
    'appro_record_id_eia': 'record_id_eia_fuzzy',
    'plant_part': 'plant_part_name_match',
    'appro_part_label': 'plant_part',
    'true_gran': 'true_gran_name_match'
}
"""dict: to rename the columns from the master unit list. Because we want
to fuzzy match with all possible MUL records names but only use the true
granualries."""

DEPRISH_COLS = [
    'utility_id_ferc1', 'utility_id_pudl', 'utility_name_ferc1', 'state',
    'plant_id_pudl', 'plant_name', 'report_year']

###############################################################################
# Prep the inputs
###############################################################################


def prep_int_ids(int_ids):
    """Prep dictionary of column names (key) with nullable int dype (value)."""
    # prep dtype arg for integer columns for read_excel
    return {i: pd.Int64Dtype() for i in int_ids}


def prep_deprish(file_path_deprish, plant_parts_df,
                 sheet_name_deprish,  key_deprish):
    """
    Prep the depreciation dataframe for use in match_merge.

    Grab the table from excel, grab only the records which can be associated
    with EIA master unit list records by plant_id_pudl, and do some light
    cleaning.
    """
    logger.info(f"Reading the depreciation data from {file_path_deprish}")
    # read in the depreciation sheet, assign types when required
    deprish_df = (pd.read_excel(
        file_path_deprish, skiprows=0, sheet_name=sheet_name_deprish,
        dtypes=prep_int_ids(INT_IDS))
        .astype({'report_date': 'datetime64[ns]',
                 'plant_id_pudl': pd.Int64Dtype()})
        .assign(report_year=lambda x: x.report_date.dt.year)
        .dropna(subset=RESTRICT_MATCH_COLS)
    )
    # because we are comparing to the EIA-based master unit list, we want to
    # only include records which are associated with plant_id_pudl's that are
    # in the master unit list.
    deprish_ids = (pd.merge(
        deprish_df,
        plant_parts_df[RESTRICT_MATCH_COLS].drop_duplicates().dropna(),
        how='outer', indicator=True, on=RESTRICT_MATCH_COLS)
    )
    deprish_ids = (
        deprish_ids.loc[deprish_ids._merge == 'both']
        .drop_duplicates(subset=['plant_name', 'report_date'])
        # there are some records in the depreciation df that have no
        # names.... so they've got to go
        .dropna(subset=['plant_name'])
    )
    deprish_ids[key_deprish] = pudl.helpers.cleanstrings_series(
        deprish_ids[key_deprish], str_map=STRINGS_TO_CLEAN)
    return deprish_ids


def prep_master_parts_eia(plant_parts_df, key_mul):
    """Prepare the EIA master plant parts."""
    plant_parts_df[key_mul] = pudl.helpers.cleanstrings_series(
        plant_parts_df[key_mul], str_map=STRINGS_TO_CLEAN)
    return plant_parts_df


###############################################################################
# Fuzzy Matching
###############################################################################


def get_plant_year_util_list(plant_name, df1, df2, key2):
    """
    Get the possible key matches from df2 a plant_id_pudl and report_year.

    This selects for the plant id and year for each of the df1 records to
    match. This is for use within `get_fuzzy_matches`.
    """
    logger.debug(plant_name)
    plant_id_pudls = df1.loc[df1.plant_name ==
                             plant_name, 'plant_id_pudl'].values
    report_years = df1.loc[df1.plant_name ==
                           plant_name, 'report_year'].values
    utility_ids = df1.loc[df1.plant_name ==
                          plant_name, 'utility_id_pudl'].values
    names = df2.loc[(df2.plant_id_pudl.isin(plant_id_pudls))
                    & (df2.report_year.isin(report_years))
                    & (df2.utility_id_pudl.isin(utility_ids))
                    ][key2].to_list()
    return names


def get_fuzzy_matches(deprish_df, mul_df, key_deprish, key_mul, threshold=75):
    """
    Get fuzzy matches on df1 using token_sort_ratio and extractOne.

    Using fuzzywuzzy's fuzzy string matching, this function matches each value
    in the key1 column with the best matched key1.

    Args:
        deprish_df (pandas.Dataframe): is the left table to join
        mul_df (pandas.Dataframe): is the right table to join
        key_deprish (str): is the key column of the left table
        key_mul (str): is the key column of the right table
        threshold (int): is how close the matches should be to return a match,
        based on Levenshtein distance. Range between 0 and 100.

    Returns:
        pandas.DataFrame
    """
    # get the best match for each valye of the key1 column
    match_tuple_series = deprish_df[key_deprish].apply(
        lambda x: process.extractOne(
            x, get_plant_year_util_list(x, deprish_df, mul_df, key_mul),
            scorer=fuzz.token_sort_ratio)
    )
    # process.extractOne returns a tuple with the matched name and the
    # match's score, so match_tuple_series contains tuples of the matching
    # plant name and the score. The plant_name_match assign only assigns the
    # match if the score is greater than or equal to the threshold.
    deprish_df = deprish_df.assign(
        matches=match_tuple_series,
        plant_name_match=match_tuple_series.apply(
            lambda x: x[0] if x[1] >= threshold else None)
    )

    matches_perct = (len(deprish_df[deprish_df.plant_name_match.notnull()])
                     / len(deprish_df))
    logger.info(f"Matches: {matches_perct:.02%}")
    return deprish_df


def match_merge(deprish_df, mul_df, key_deprish, key_mul):
    """Generate fuzzy matches and merge relevant colums from eia."""
    logger.info("Merging fuzzy matches.")
    # we are going to match with all of the names from the
    # master unit list, but then use the "true granularity"
    match_merge_df = (pd.merge(
        get_fuzzy_matches(
            deprish_df, mul_df,
            key_deprish=key_deprish, key_mul=key_mul,
            threshold=75)[DEPRISH_COLS + ['plant_name_match']],
        mul_df.reset_index().drop_duplicates(
            subset=['report_year', 'plant_name_new'])[MUL_COLS],
        left_on=['report_year', 'plant_name_match'],
        right_on=['report_year', key_mul], how='left')
        # rename the ids so that we have the "true granularity"
        # Every MUL record has identifying columns for it's true granualry,
        # even when the true granularity is the same record, so we can use the
        # true gran columns across the board.
        .rename(columns=MUL_RENAME)
    )
    logger.info(f"Matching resulted in {len(match_merge_df)} connections.")
    return match_merge_df


def add_overrides(deprish_match, file_path_deprish, sheet_name_output):
    """
    Add the overrides into the matched depreciation records.

    Args:
        deprish_match (pandas.DataFrame):
        file_path_deprish (os.PathLike): path to the excel workbook which
           contains depreciation data.
       sheet_name_output (string): name of the excel tab which the matches
          will be output.

    Returns:
        pandas.DataFrame: augmented version of deprish_match with overrides.
    """
    try:
        overrides_df = (
            pd.read_excel(
                file_path_deprish, skiprows=0, sheet_name=sheet_name_output,)
            .drop(columns=['record_id_eia', 'record_id_eia_fuzzy'])
        )
        logger.info(f"Adding overrides from {sheet_name_output}.")
        # concat, sort so the True overrides are at the top and drop dupes
        deprish_match_full = (
            pd.concat([deprish_match, overrides_df])
            .sort_values('record_id_eia_override')
            .drop_duplicates(subset=DEPRISH_COLS, keep='first')
            .assign(record_id_eia=lambda x:
                    x.record_id_eia_override.fillna(x.record_id_eia_fuzzy))
        )
    except XLRDError:
        logger.info(f"No sheet {sheet_name_output}, so no overrides added.")
        deprish_match_full = deprish_match
    return deprish_match_full


def match_deprish_eia(file_path_mul, file_path_deprish,
                      sheet_name_deprish, sheet_name_output):
    """Prepare the depreciation and master unit list and match on name cols."""
    key_deprish = 'plant_name'
    key_mul = 'plant_name_new'
    logger.info('Grab or generate master unit list.')
    plant_parts_df = (
        make_plant_parts_eia.get_master_unit_list_eia(file_path_mul)
        .reset_index()
    )
    deprish_df = prep_deprish(
        file_path_deprish,
        plant_parts_df,
        sheet_name_deprish=sheet_name_deprish,
        key_deprish=key_deprish
    )
    mul_df = prep_master_parts_eia(plant_parts_df, key_mul=key_mul)
    deprish_match = (
        match_merge(deprish_df, mul_df,
                    key_deprish=key_deprish, key_mul=key_mul)
        .pipe(add_overrides, file_path_deprish=file_path_deprish,
              sheet_name_output=sheet_name_output)
        .pipe(pudl.helpers.organize_cols,
              # we want to pull the used columns to the front, but there is
              # some overlap in columns from these two datasets. And we have
              # renamed some of the columns from the master unit list.
              list(set(DEPRISH_COLS + [MUL_RENAME.get(c, c)
                                       for c in MUL_COLS])))
    )
    possible_matches_mul = (
        pd.merge(
            plant_parts_df.dropna(subset=RESTRICT_MATCH_COLS),
            deprish_df[RESTRICT_MATCH_COLS].drop_duplicates(),
            on=RESTRICT_MATCH_COLS)
        .pipe(pudl.helpers.organize_cols, MUL_COLS)
    )
    return deprish_match, possible_matches_mul

###############################################################################
# EXPORT
###############################################################################


def generate_depreciation_matches(file_path_mul,
                                  file_path_deprish,
                                  sheet_name_deprish,
                                  sheet_name_output):
    """
    Generate the matched names and save to execl.

    Args:
         file_path_mul (pathlib.Path): path to the master unit list.
         file_path_deprish (os.PathLike): path to the excel workbook which
            contains depreciation data.
         sheet_name_deprish (str): name of the excel tab which contains the
            plant names from the depreciation data.
         sheet_name_output (string): name of the excel tab which the matches
            will be output.

    Returns:
        pandas.DataFrame : dataframe including matched names from depreciation
            data to names in the mater unit list, including appropirate id's
            from the master unit list.
    """
    if not file_path_deprish.is_file():
        raise FileNotFoundError(
            f"File does not exist: {file_path_deprish}"
            "Depretiation file must exist."
        )
    deprish_match_df, possible_matches_mul_df = match_deprish_eia(
        file_path_mul, file_path_deprish,
        sheet_name_deprish=sheet_name_deprish,
        sheet_name_output=sheet_name_output
    )
    sheets_df_dict = {sheet_name_output: deprish_match_df,
                      "Subset of Master Unit List": possible_matches_mul_df}
    save_to_workbook(file_path=file_path_deprish,
                     sheets_df_dict=sheets_df_dict)
    return deprish_match_df


def save_to_workbook(file_path, sheets_df_dict):
    """
    Save dataframes to sheets in an existing workbook.

    Args:
        file_path (path-like): the location of the excel workbook.
        sheets_df_dict (dict): dictionary of the names of sheets (keys) of
            where their corresponding dataframes (values) should end up in the
            workbook.
    """
    logger.info(f"Saving dataframe to {file_path}")
    if not file_path.exists():
        raise AssertionError(f'file path {file_path} does not exist')
    workbook1 = load_workbook(file_path)
    writer = pd.ExcelWriter(file_path, engine='openpyxl')
    writer.book = workbook1
    for sheet_name, df in sheets_df_dict.items():
        if sheet_name in workbook1.sheetnames:
            logger.info(f"Removing {sheet_name} from {file_path}")
            workbook1.remove(workbook1[sheet_name])
            if sheet_name in workbook1.sheetnames:
                raise AssertionError(f"{sheet_name} was not removed")
        df.to_excel(writer, sheet_name=sheet_name, index=False)

    writer.save()
    writer.close()

###############################################################################
# Command line script
###############################################################################


def parse_command_line(argv):
    """
    Parse script command line arguments. See the -h option.

    Args:
        argv (list): command line arguments including caller file name.

    Returns:
        dict: A dictionary mapping command line arguments to their values.

    """
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        'file_path_deprish',
        type=str,
        help='path to the excel workbook which contains depreciation data.')
    parser.add_argument(
        '--file_path_mul',
        default=pathlib.Path('master_unit_list.pkl.gz'),
        type=str,
        help='path to the master unit list. Default: master_unit_list.csv.gz')
    parser.add_argument(
        '--sheet_name_deprish',
        default='Depreciation Studies Raw',
        type=str,
        help="""name of the excel tab which contains the plant names from the
        depreciation data. Default: Depreciation Studies Raw""")
    parser.add_argument(
        '--sheet_name_output',
        default='EIA to depreciation matches',
        type=str,
        help="""name of the excel tab which the matches will be output.
        Default: EIA to depreciation matches""")
    arguments = parser.parse_args(argv[1:])
    return arguments


def main():
    """Match depreciation and master unit list records. Save to excel."""
    args = parse_command_line(sys.argv)

    _ = generate_depreciation_matches(
        file_path_mul=pathlib.Path(args.file_path_mul),
        file_path_deprish=pathlib.Path(args.file_path_deprish),
        sheet_name_deprish=args.sheet_name_deprish,
        sheet_name_output=args.sheet_name_output)


if __name__ == '__main__':
    sys.exit(main())
