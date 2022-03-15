"""
Connect the EIA plant-parts with depreciation records from PUCs and FERC1.

This module connects the records from the depreciation data to their
appropirate ids in the EIA plant-parts. The plant-parts is generated
via `make_plant_parts_eia.py`; see the documenation there for additional
details about the plant-parts. The depreciation data is annual
depreciation studies from PUC and FERC1 data that have been compiled into an
excel spreadsheet. The plant-parts is a compilation of various slices of
EIA records.
"""

import argparse
import logging
import sys

import pandas as pd
import pudl
import sqlalchemy as sa
from fuzzywuzzy import fuzz, process
from openpyxl import load_workbook
from xlrd import XLRDError

import pudl_rmi
import pudl_rmi.make_plant_parts_eia as make_plant_parts_eia

logger = logging.getLogger(__name__)

STRINGS_TO_CLEAN = {
    "combined cycle": ["CC"],
    "combustion turbine": ["CT"],
}

RESTRICT_MATCH_COLS = ['plant_id_eia', 'utility_id_pudl', 'report_year']

PPE_COLS = [
    'record_id_eia', 'plant_name_new', 'plant_part', 'report_year',
    'report_date', 'ownership', 'plant_name_eia', 'plant_id_eia',
    'generator_id', 'unit_id_pudl', 'prime_mover_code', 'energy_source_code_1',
    'technology_description', 'ferc_acct_name', 'utility_id_eia',
    'utility_id_pudl', 'true_gran', 'appro_part_label', 'appro_record_id_eia',
    'record_count', 'fraction_owned', 'ownership_dupe', 'operational_status',
    'operational_status_pudl'
]

PPE_RENAME = {
    # 'record_id_eia': 'record_id_eia_name_match',
    # this rename doesn't have huge downstream implications, but i don't know
    # if this first rename is correct.... (if it is determined to be unecessary
    # and removed, we also need to remove a .drop(columns in match_deprish_eia)
    'appro_record_id_eia': 'record_id_eia_fuzzy',
    'plant_part': 'plant_part_name_match',
    'appro_part_label': 'plant_part',
    'true_gran': 'true_gran_name_match'
}
"""dict: to rename the columns from the EIA plant-parts. Because we want
to fuzzy match with all possible PPE records names but only use the true
granualries."""

IDX_DEPRISH_COLS = [
    'utility_id_ferc1', 'utility_id_pudl', 'utility_name_ferc1',
    'plant_id_eia', 'plant_part_name', 'report_year']

###############################################################################
# Prep the inputs
###############################################################################


def prep_deprish(deprish, plant_parts_eia, key_deprish):
    """
    Prep the depreciation dataframe for use in match_merge.

    Grab only the records which can be associated with EIA plant-parts records
    by plant_id_pudl, and do some light cleaning.
    """
    # we could grab the output here instead of the input file...
    deprish = (
        deprish.assign(report_year=lambda x: x.report_date.dt.year)
        .convert_dtypes(convert_floating=False)
    )

    deprish.loc[:, key_deprish] = pudl.helpers.cleanstrings_series(
        deprish.loc[:, key_deprish], str_map=STRINGS_TO_CLEAN)
    # because we are comparing to the EIA-based plant-parts, we want to
    # only include records which are associated with plant_id_pudl's that are
    # in the EIA plant-parts.
    deprish_ids = pd.merge(
        deprish,
        plant_parts_eia[RESTRICT_MATCH_COLS]
        .drop_duplicates().astype({'report_year': pd.Int64Dtype()}),
        how='outer',
        indicator=True,
        on=RESTRICT_MATCH_COLS,
        validate='m:1'
    )
    # check the number of depreciation records that should have EIA plant-part
    # matches.
    # TODO: go through all of these to reassign plant_id_eia!!! and turn down
    # the acceptable number of baddies
    baddies = (
        deprish_ids.loc[(deprish_ids._merge != 'both')]
        .dropna(subset=RESTRICT_MATCH_COLS + ['plant_part_name'])
        .drop_duplicates(subset=['plant_part_name'])
    )
    if len(baddies) > 270:
        raise AssertionError(
            f"Found {len(baddies)} depreciation records which don't have "
            "cooresponding EIA plant-parts records. Check plant_id_eia's "
            f"in {pudl_rmi.DEPRISH_RAW_XLSX}"
        )
    deprish_ids = (
        deprish_ids.loc[deprish_ids._merge == 'both']
        # there are some records in the depreciation df that have no
        # names.... so they've got to go
        .dropna(subset=['plant_part_name', '_merge'])
        .drop(columns=['_merge'])
        .convert_dtypes(convert_floating=False)
    )
    return deprish_ids


def prep_master_parts_eia(plant_parts_df, deprish, key_ppe):
    """Prepare the EIA master plant parts."""
    # restrict the possible matches to only those that match on the
    # RESTRICT_MATCH_COLS
    options_index = (deprish[RESTRICT_MATCH_COLS].drop_duplicates()
                     .set_index(RESTRICT_MATCH_COLS).index)
    plant_parts_df = (
        plant_parts_df
        .reset_index()  # convert the record_id_eia index to a column
        .set_index(RESTRICT_MATCH_COLS).loc[options_index]
        .reset_index()
        .convert_dtypes(convert_floating=False)
    )
    plant_parts_df.loc[:, key_ppe] = pudl.helpers.cleanstrings_series(
        plant_parts_df[key_ppe], str_map=STRINGS_TO_CLEAN)
    return plant_parts_df


###############################################################################
# Fuzzy Matching
###############################################################################


def get_plant_year_util_list(plant_name, deprish, ppe, key_ppe):
    """
    Get the possible key matches from df2 a plant_id_pudl and report_year.

    This selects for the plant id and year for each of the df1 records to
    match. This is for use within `get_fuzzy_matches`.
    """
    logger.debug(plant_name)
    options_index = (
        deprish.loc[deprish.plant_part_name ==
                    plant_name, RESTRICT_MATCH_COLS]
        .drop_duplicates().set_index(RESTRICT_MATCH_COLS).index)

    # get the set of possible names
    names = (ppe.set_index(RESTRICT_MATCH_COLS)
             .loc[options_index, key_ppe].to_list())
    return names


def get_fuzzy_matches(deprish, ppe, key_deprish, key_ppe, threshold=75):
    """
    Get fuzzy matches on df1 using token_sort_ratio and extractOne.

    Using fuzzywuzzy's fuzzy string matching, this function matches each value
    in the key1 column with the best matched key1.

    Args:
        deprish (pandas.Dataframe): is the left table to join
        ppe (pandas.Dataframe): is the right table to join
        key_deprish (str): is the key column of the left table
        key_ppe (str): is the key column of the right table
        threshold (int): is how close the matches should be to return a match,
            based on Levenshtein distance. Range between 0 and 100.

    Returns:
        pandas.DataFrame
    """
    logger.info("Generating fuzzy matches.")
    # get the best match for each valye of the key1 column
    match_tuple_series = deprish[key_deprish].apply(
        lambda x: process.extractOne(
            x, get_plant_year_util_list(x, deprish, ppe, key_ppe),
            scorer=fuzz.token_sort_ratio)
    )
    # process.extractOne returns a tuple with the matched name and the
    # match's score, so match_tuple_series contains tuples of the matching
    # plant name and the score. The plant_name_match assign only assigns the
    # match if the score is greater than or equal to the threshold.
    deprish = deprish.assign(
        matches=match_tuple_series,
        plant_name_match=match_tuple_series.apply(
            lambda x: x[0] if x[1] >= threshold else None)
    )

    matches_perct = (len(deprish[deprish.plant_name_match.notnull()])
                     / len(deprish))
    logger.info(f"Matches: {matches_perct:.02%}")
    logger.info(f"Matching resulted in {len(deprish)} connections.")
    return deprish


def add_overrides(deprish_match, file_path_deprish, sheet_name_output):
    """
    Add the overrides into the matched depreciation records.

    Args:
        deprish_match (pandas.DataFrame):
        file_path_deprish (os.PathLike): path to the excel workbook which
           contains depreciation data.
       sheet_name_output (string): name of the excel tab which the matches
          will be output.

    Returns:
        pandas.DataFrame: augmented version of deprish_match with overrides.
    """
    try:
        overrides_df = (
            pd.read_excel(
                file_path_deprish, skiprows=0, sheet_name=sheet_name_output,)
        )
        overrides_df = (
            overrides_df[
                overrides_df.filter(like='record_id_eia_override')
                .notnull().any(axis='columns')
            ]
            [RESTRICT_MATCH_COLS + ['plant_part_name', 'data_source'] +
             list(overrides_df.filter(like='record_id_eia_override').columns)])
        logger.info(
            f"Adding {len(overrides_df)} overrides from {sheet_name_output}.")
        # concat, sort so the True overrides are at the top and drop dupes
        deprish_match_full = (
            pd.merge(
                deprish_match.convert_dtypes(convert_floating=False),
                overrides_df.convert_dtypes(convert_floating=False),
                on=RESTRICT_MATCH_COLS + ['plant_part_name', 'data_source'],
                how='outer'
            )
            .drop_duplicates()
            .assign(record_id_eia=lambda x:
                    x.record_id_eia_override.fillna(x.record_id_eia_fuzzy))
        )
    except XLRDError:
        logger.info(f"No sheet {sheet_name_output}, so no overrides added.")
        deprish_match_full = deprish_match
    return deprish_match_full


def match_deprish_eia(deprish, plant_parts_eia, sheet_name_output):
    """Prepare the depreciation and EIA plant-parts and match on name cols."""
    key_deprish = 'plant_part_name'
    key_ppe = 'plant_name_new'
    deprish = prep_deprish(
        deprish=deprish,
        plant_parts_eia=plant_parts_eia,
        key_deprish=key_deprish
    )
    ppe = prep_master_parts_eia(plant_parts_eia, deprish, key_ppe=key_ppe)
    deprish_match = (
        get_fuzzy_matches(
            deprish=deprish, ppe=ppe,
            key_deprish=key_deprish, key_ppe=key_ppe,
            threshold=75)
        .pipe(add_record_id_fuzzy, plant_parts_eia=ppe, key_ppe=key_ppe)
        .pipe(add_overrides, file_path_deprish=pudl_rmi.DEPRISH_RAW_XLSX,
              sheet_name_output=sheet_name_output)
    )
    # merge in the rest of the ppe columns
    # we want to have the columns from the PPE, not from the fuzzy merged
    # option. so we're dropping all of the shared columns before merging
    ppe_non_id_cols = [c for c in PPE_COLS if c != 'record_id_eia']
    deprish_match = (
        deprish_match
        .drop(columns=[c for c in deprish_match
                       if c in ppe_non_id_cols
                       and c not in RESTRICT_MATCH_COLS])
        .merge(
            ppe.reset_index()[PPE_COLS],
            on=['record_id_eia'] + RESTRICT_MATCH_COLS,
            how='left',
            validate='m:1',
        )
        # rename the ids so that we have the "true granularity"
        # Every PPE record has identifying columns for it's true granualry,
        # even when the true granularity is the same record, so we can use the
        # true gran columns across the board.
        # we first want to remove this fuzzy id before we rename the
        # appropirate ID column so we don't have duplicate columns
        .drop(columns=['record_id_eia_fuzzy'])
        .rename(columns=PPE_RENAME)
        # reassign_id_ownership_dupes will fail with nulls in this bool col
        .assign(ownership_dupe=lambda x: x.ownership_dupe.fillna(False))
        .pipe(make_plant_parts_eia.reassign_id_ownership_dupes)
        .convert_dtypes(convert_floating=False)

    )
    first_cols = [
        'plant_part_name', 'utility_name_ferc1', 'report_year',
        'plant_name_match', 'record_id_eia', 'record_id_eia_fuzzy',
    ] + list(deprish_match.filter(like='_override'))
    deprish_match = deprish_match.loc[
        :, first_cols + [x for x in deprish_match.columns
                         if x not in first_cols]]
    return deprish_match


def add_record_id_fuzzy(deprish, plant_parts_eia, key_ppe):
    """Merge in relevant columns from EIA plant-parts."""
    left_on = [
        'report_year',
        'utility_id_pudl',
        'plant_name_match',
        'plant_id_eia']
    right_on = [
        'report_year',
        'utility_id_pudl',
        key_ppe,
        'plant_id_eia']
    # we're adding the appro ID and then we'll reassign that column as
    # record_id_eia_fuzzy below
    match_merge_df = (
        pd.merge(
            deprish,
            plant_parts_eia
            .drop_duplicates(subset=['report_year', 'plant_name_new'])
            [right_on + ['appro_record_id_eia']],
            left_on=left_on,
            right_on=right_on,
            how='left',
            validate='m:1'
        )
        .rename(columns={'appro_record_id_eia': 'record_id_eia_fuzzy'})
    )
    return match_merge_df


def grab_possible_plant_part_eia_matches(plant_parts_eia, deprish):
    """
    Get the part of the EIA plant-parts that could match with depreciation.

    Returns:
        pandas.DataFrame: A subset of the EIA plant-parts that cooresponds to
        possible matches for the depreciation data based on the
        ``RESTRICT_MATCH_COLS``.
    """
    possible_matches_ppe = (
        pd.merge(
            plant_parts_eia.reset_index()
            .convert_dtypes(convert_floating=False)
            .dropna(subset=RESTRICT_MATCH_COLS),
            deprish[RESTRICT_MATCH_COLS].drop_duplicates(),
            on=RESTRICT_MATCH_COLS)
        .pipe(pudl.helpers.organize_cols, PPE_COLS)
    )
    return possible_matches_ppe

###############################################################################
# EXPORT
###############################################################################


def execute(
    deprish,
    plant_parts_eia,
    sheet_name_output='EIA to depreciation matches',
    save_to_xls=True,
):
    """
    Generate the matched names and save to excel.

    This method generates a link between depreciation records and the EIA
    plant-parts. It generates all of the options that could have been matched
    from the EIA plant-parts; this will help with generating mannual
    overrides. It then saves these outputs into the same spreadsheet that the
    depreciation records were pulled from.

    Args:
        plant_parts_eia (panda.DataFrame): EIA plant-parts - table of
            "plant-parts" which are groups of aggregated EIA generators
            that coorespond to portions of plants from generators to fuel
            types to whole plants.
        sheet_name_output (string): name of the excel tab which the matches
            will be output and the place where we will grab the overrides for
            the matches. Default is: 'EIA to depreciation matches'.
        save_to_xls (boolean): If True, save the outputs to the workbook.
            Default is True. If False, the outputs are not saved - it reduces
            time and the dataframe is still returned.

    Returns:
        pandas.DataFrame : dataframe including matched names from depreciation
            data to names in the EIA plant-parts, including appropirate id's
            from the EIA plant-parts.
    """
    deprish_match = match_deprish_eia(
        deprish,
        plant_parts_eia,
        sheet_name_output=sheet_name_output
    )
    possible_matches_ppe = grab_possible_plant_part_eia_matches(
        plant_parts_eia, deprish_match
    )
    if save_to_xls:
        sheets_df_dict = {
            sheet_name_output: deprish_match,
            "Subset of Master Unit List": possible_matches_ppe}
        save_to_workbook(file_path=pudl_rmi.DEPRISH_RAW_XLSX,
                         sheets_df_dict=sheets_df_dict)
    return deprish_match


def save_to_workbook(file_path, sheets_df_dict):
    """
    Save dataframes to sheets in an existing workbook.

    This method enables us to save multiple dataframes into differnt tabs in
    the same excel workbook. If those tabs already exist, the default process
    is to make save a new tab with a suffix, so we remove the tab if it exists
    before saving.

    Args:
        file_path (pathlib.Path): the location of the excel workbook.
        sheets_df_dict (dict): dictionary of the names of sheets (keys) of
            where their corresponding dataframes (values) should end up in the
            workbook.
    """
    logger.info(f"Saving dataframe to {file_path}")
    if not file_path.exists():
        raise AssertionError(f'file path {file_path} does not exist')
    workbook1 = load_workbook(file_path)
    writer = pd.ExcelWriter(file_path, engine='openpyxl')
    writer.book = workbook1
    for sheet_name, df in sheets_df_dict.items():
        if sheet_name in workbook1.sheetnames:
            logger.info(f"Removing {sheet_name} from {file_path}")
            workbook1.remove(workbook1[sheet_name])
            if sheet_name in workbook1.sheetnames:
                raise AssertionError(f"{sheet_name} was not removed")
        df.to_excel(writer, sheet_name=sheet_name, index=False)

    writer.save()
    writer.close()

###############################################################################
# Command line script
###############################################################################


def parse_command_line(argv):
    """
    Parse script command line arguments. See the -h option.

    Args:
        argv (list): command line arguments including caller file name.

    Returns:
        dict: A dictionary mapping command line arguments to their values.

    """
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        'clobber',
        '-c',
        '--clobber',
        action='store_true',
        default=False,
        help="Clobber existing depreciation/EIA pickled output if it exists.")
    parser.add_argument(
        'clobber_deprish',
        action='store_true',
        default=False,
        help="Clobber existing interim pickled output of depreciation data.")
    parser.add_argument(
        'clobber_plant_parts_eia',
        action='store_true',
        default=False,
        help="Clobber existing interim pickled EIA plant-parts output.")
    arguments = parser.parse_args(argv[1:])
    return arguments


def main():
    """Match depreciation and EIA plant-parts records. Save to excel."""
    args = parse_command_line(sys.argv)

    pudl_settings = pudl.workspace.setup.get_defaults()
    pudl_engine = sa.create_engine(pudl_settings["pudl_db"])

    pudl_out = pudl.output.pudltabl.PudlTabl(
        pudl_engine,
        freq='AS',
        fill_fuel_cost=True,
        roll_fuel_cost=True,
        fill_net_gen=False
    )

    rmi_out = pudl_rmi.coordinate.Output(pudl_out)

    _ = rmi_out.grab_deprish_to_eia(
        clobber=args.clobber,
        clobber_deprish=args.clobber_deprish,
        clobber_plant_parts_eia=args.clobber_plant_parts_eia,
    )


if __name__ == '__main__':
    sys.exit(main())
